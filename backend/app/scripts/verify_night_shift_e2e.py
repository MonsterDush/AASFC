from __future__ import annotations

import json
import os
from calendar import monthrange
from datetime import date, time
from io import BytesIO

from fastapi import BackgroundTasks
from openpyxl import load_workbook
from sqlalchemy import inspect, select, text
from sqlalchemy.exc import IntegrityError

from app.core.db import SessionLocal
from app.models import (
    DailyReport,
    DailyReportTipAllocation,
    Department,
    Expense,
    ExpenseCategory,
    ExpenseRecognitionEntry,
    FinanceEntry,
    NotificationJob,
    PaymentMethod,
    Shift,
    ShiftInterval,
    User,
    Venue,
    VenueMember,
    VenuePosition,
)
from app.routers import venue_reports, venue_revenue_exports, venue_shifts
from app.schemas.venue_reports import (
    DailyReportCloseIn,
    DailyReportUpsertIn,
    ReportValueIn,
)
from app.schemas.venue_shifts import ShiftAssignmentAddIn, ShiftCreateIn
from app.scripts.bootstrap_e2e_data import (
    DEFAULT_VENUE_NAME,
    require_safe_e2e_database,
)
from app.services.finance.summary import get_day_finance_summary
from app.services.finance.expenses import rebuild_expense_allocations_for_expense
from app.services.payroll.component_calculations import interval_duration_minutes
from app.services.payroll.day_breakdown import build_member_day_breakdown
from app.services.xlsx_export import build_revenue_xlsx
from app.settings import settings


CHECK_CONSTRAINTS = {
    "shifts": "ck_shifts_shift_slot_valid",
    "daily_reports": "ck_daily_reports_shift_slot_valid",
    "daily_report_attachments": "ck_daily_report_attachments_shift_slot_valid",
    "shift_schedule_template_items": "ck_shift_schedule_template_items_shift_slot_valid",
    "expenses": "ck_expenses_shift_slot_valid",
    "recurring_expense_rules": "ck_recurring_expense_rules_shift_slot_valid",
    "expense_recognition_entries": "ck_expense_recognition_entries_shift_slot_valid",
}
OVERNIGHT_INTERVAL_TITLE = "E2E ночь 22:00–04:00"


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _month_boundary() -> tuple[date, date]:
    today = date.today()
    last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
    if today.month == 12:
        next_month = date(today.year + 1, 1, 1)
    else:
        next_month = date(today.year, today.month + 1, 1)
    return last_day, next_month


def _load_fixture_people(db):
    venue_name = str(os.getenv("E2E_VENUE_NAME") or DEFAULT_VENUE_NAME).strip()
    venue = db.execute(select(Venue).where(Venue.name == venue_name).order_by(Venue.id.asc())).scalar_one_or_none()
    _assert(venue is not None, "E2E venue is missing; run the seed first")

    owner = db.execute(
        select(User)
        .join(VenueMember, VenueMember.user_id == User.id)
        .where(
            VenueMember.venue_id == int(venue.id),
            VenueMember.venue_role == "OWNER",
            VenueMember.is_active.is_(True),
        )
        .order_by(User.id.asc())
    ).scalar_one_or_none()
    _assert(owner is not None, "E2E owner is missing")

    staff_row = db.execute(
        select(User, VenuePosition)
        .join(VenueMember, VenueMember.user_id == User.id)
        .join(
            VenuePosition,
            (VenuePosition.member_user_id == User.id) & (VenuePosition.venue_id == int(venue.id)),
        )
        .where(
            VenueMember.venue_id == int(venue.id),
            VenueMember.venue_role == "STAFF",
            VenueMember.is_active.is_(True),
            VenuePosition.is_active.is_(True),
        )
        .order_by(User.id.asc())
    ).first()
    _assert(staff_row is not None, "E2E staff position is missing")
    return venue, owner, staff_row[0], staff_row[1]


def _assert_database_constraints(db) -> None:
    inspector = inspect(db.get_bind())
    for table_name, constraint_name in CHECK_CONSTRAINTS.items():
        names = {str(item.get("name")) for item in inspector.get_check_constraints(table_name) if item.get("name")}
        _assert(
            constraint_name in names,
            f"{constraint_name} is missing from {table_name}",
        )


def _prepare_catalog_values(db, *, venue_id: int) -> tuple[int, int]:
    payment_method_id = db.execute(
        select(PaymentMethod.id)
        .where(
            PaymentMethod.venue_id == int(venue_id),
            PaymentMethod.is_active.is_(True),
        )
        .order_by(PaymentMethod.id.asc())
        .limit(1)
    ).scalar_one()
    department_id = db.execute(
        select(Department.id)
        .where(
            Department.venue_id == int(venue_id),
            Department.is_active.is_(True),
        )
        .order_by(Department.id.asc())
        .limit(1)
    ).scalar_one()
    return int(payment_method_id), int(department_id)


def _create_overnight_interval(db, *, venue_id: int) -> ShiftInterval:
    interval = db.execute(
        select(ShiftInterval).where(
            ShiftInterval.venue_id == int(venue_id),
            ShiftInterval.title == OVERNIGHT_INTERVAL_TITLE,
        )
    ).scalar_one_or_none()
    if interval is None:
        interval = ShiftInterval(
            venue_id=int(venue_id),
            title=OVERNIGHT_INTERVAL_TITLE,
            start_time=time(22, 0),
            end_time=time(4, 0),
            is_active=True,
        )
        db.add(interval)
    else:
        interval.start_time = time(22, 0)
        interval.end_time = time(4, 0)
        interval.is_active = True
    db.commit()
    db.refresh(interval)
    return interval


def _assert_invalid_slot_rejected(db, *, shift_id: int) -> None:
    rejected = False
    try:
        with db.begin_nested():
            db.execute(
                text("UPDATE shifts SET shift_slot = 'INVALID' WHERE id = :shift_id"),
                {"shift_id": int(shift_id)},
            )
    except IntegrityError:
        rejected = True
    _assert(rejected, "PostgreSQL accepted an invalid shift_slot")


def _prepare_common_expense_allocation(
    db,
    *,
    venue_id: int,
    target_date: date,
    owner_user_id: int,
) -> tuple[Expense, list[ExpenseRecognitionEntry]]:
    category_id = db.execute(
        select(ExpenseCategory.id)
        .where(ExpenseCategory.venue_id == int(venue_id))
        .order_by(ExpenseCategory.id.asc())
        .limit(1)
    ).scalar_one()
    marker = f"E2E common shift allocation {target_date.strftime('%Y-%m')}"
    expense = db.execute(
        select(Expense).where(
            Expense.venue_id == int(venue_id),
            Expense.comment == marker,
        )
    ).scalar_one_or_none()
    recognition_count = monthrange(target_date.year, target_date.month)[1] * 2
    amount_minor = recognition_count * 100 + 1
    if expense is None:
        expense = Expense(
            venue_id=int(venue_id),
            category_id=int(category_id),
            amount_minor=amount_minor,
            expense_date=target_date,
            shift_slot="TOTAL",
            spread_months=1,
            status="CONFIRMED",
            comment=marker,
            created_by_user_id=int(owner_user_id),
        )
        db.add(expense)
        db.flush()
    else:
        expense.category_id = int(category_id)
        expense.amount_minor = amount_minor
        expense.expense_date = target_date
        expense.shift_slot = "TOTAL"
        expense.spread_months = 1
        expense.status = "CONFIRMED"

    rebuild_expense_allocations_for_expense(db=db, expense=expense)
    db.commit()
    rows = (
        db.execute(
            select(ExpenseRecognitionEntry)
            .where(ExpenseRecognitionEntry.expense_id == int(expense.id))
            .order_by(
                ExpenseRecognitionEntry.recognition_date.asc(),
                ExpenseRecognitionEntry.shift_slot.asc(),
            )
        )
        .scalars()
        .all()
    )
    _assert(len(rows) == recognition_count, "Common expense was not split into all month shift slots")
    _assert(
        sum(int(row.amount_minor or 0) for row in rows) == amount_minor, "Shift expense split changed the monthly total"
    )
    _assert(
        {str(row.shift_slot) for row in rows} == {"DAY", "NIGHT"},
        "Common expense recognition is missing DAY or NIGHT",
    )
    return expense, list(rows)


def _build_and_assert_xlsx(
    *,
    venue_name: str,
    report_id: int,
    report_rows: list[dict],
    value_rows: list[dict],
) -> int:
    workbook_bytes = build_revenue_xlsx(
        month=date.today().strftime("%Y-%m"),
        mode="PAYMENTS",
        venue_name=venue_name,
        rows=[],
        total=0,
        closed_reports=len(report_rows),
        report_rows=report_rows,
        value_rows=value_rows,
    )
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    report_sheet = workbook["Отчёты"]
    matching_rows = [
        row for row in report_sheet.iter_rows(min_row=4, values_only=True) if int(row[2] or 0) == int(report_id)
    ]
    _assert(len(matching_rows) == 1, "Night report is missing from the XLSX report sheet")
    _assert(matching_rows[0][1] == "NIGHT", "XLSX lost the NIGHT slot")
    return len(workbook_bytes)


def verify_night_shift_boundary() -> dict[str, object]:
    require_safe_e2e_database(
        settings.database_url,
        confirmation=os.getenv("AXELIO_E2E_ALLOW_SEED"),
    )
    boundary_date, next_month_start = _month_boundary()
    revenue_units = 12_345
    tips_units = 1_234

    with SessionLocal() as db:
        _assert_database_constraints(db)
        venue, owner, staff, staff_position = _load_fixture_people(db)
        venue.night_shifts_enabled = True
        venue.tips_enabled = True
        venue.tips_split_mode = "EQUAL"
        db.commit()

        interval = _create_overnight_interval(db, venue_id=int(venue.id))
        _assert(
            interval_duration_minutes(interval.start_time, interval.end_time) == 360,
            "22:00–04:00 must be a six-hour interval",
        )
        common_expense, common_expense_rows = _prepare_common_expense_allocation(
            db,
            venue_id=int(venue.id),
            target_date=boundary_date,
            owner_user_id=int(owner.id),
        )

        created = venue_shifts.create_shift(
            venue_id=int(venue.id),
            payload=ShiftCreateIn(
                date=boundary_date,
                interval_id=int(interval.id),
                shift_slot="NIGHT",
            ),
            db=db,
            user=owner,
        )
        shift_id = int(created["id"])
        venue_shifts.add_shift_assignment(
            venue_id=int(venue.id),
            shift_id=shift_id,
            payload=ShiftAssignmentAddIn(venue_position_id=int(staff_position.id)),
            db=db,
            user=owner,
        )

        payment_method_id, department_id = _prepare_catalog_values(
            db,
            venue_id=int(venue.id),
        )
        report_result = venue_reports.upsert_daily_report(
            venue_id=int(venue.id),
            payload=DailyReportUpsertIn(
                date=boundary_date,
                revenue_total=revenue_units,
                tips_total=tips_units,
                payments=[
                    ReportValueIn(ref_id=payment_method_id, value=revenue_units),
                ],
                departments=[
                    ReportValueIn(ref_id=department_id, value=revenue_units),
                ],
                comment="E2E: ночь на границе месяца",
            ),
            shift_slot="NIGHT",
            db=db,
            user=owner,
        )
        report_id = int(report_result["id"])

        background_tasks = BackgroundTasks()
        close_result = venue_reports.close_daily_report(
            venue_id=int(venue.id),
            report_date=boundary_date,
            payload=DailyReportCloseIn(comment=None),
            background_tasks=background_tasks,
            shift_slot="NIGHT",
            db=db,
            user=owner,
        )
        _assert(close_result["status"] == "CLOSED", "Night report did not close")
        _assert(len(background_tasks.tasks) == 1, "Notification worker was not scheduled")

        shift = db.get(Shift, shift_id)
        report = db.get(DailyReport, report_id)
        _assert(shift is not None and shift.date == boundary_date, "Shift date crossed the month boundary")
        _assert(shift.shift_slot == "NIGHT", "Shift lost its NIGHT slot")
        _assert(report is not None and report.date == boundary_date, "Report date crossed the month boundary")
        _assert(report.status == "CLOSED" and report.shift_slot == "NIGHT", "Night report state is wrong")

        july_rows = venue_reports.list_daily_reports(
            venue_id=int(venue.id),
            month=boundary_date.strftime("%Y-%m"),
            shift_slot="NIGHT",
            db=db,
            user=owner,
        )
        august_rows = venue_reports.list_daily_reports(
            venue_id=int(venue.id),
            month=next_month_start.strftime("%Y-%m"),
            shift_slot="NIGHT",
            db=db,
            user=owner,
        )
        _assert(
            any(int(item["id"]) == report_id for item in july_rows),
            "Night report is missing from its start month",
        )
        _assert(
            all(int(item["id"]) != report_id for item in august_rows),
            "Night report leaked into the month when it ended",
        )

        finance = get_day_finance_summary(
            db=db,
            venue_id=int(venue.id),
            target_date=boundary_date,
            income_mode="PAYMENTS",
            shift_slot="NIGHT",
        )
        _assert(
            int(finance["revenue_minor"]) == revenue_units * 100,
            "Night revenue was not isolated by slot",
        )
        night_expense_minor = sum(
            int(row.amount_minor or 0)
            for row in common_expense_rows
            if row.recognition_date == boundary_date and row.shift_slot == "NIGHT"
        )
        expected_night_expense_minor = sum(
            int(amount_minor or 0)
            for amount_minor in db.execute(
                select(ExpenseRecognitionEntry.amount_minor)
                .join(Expense, Expense.id == ExpenseRecognitionEntry.expense_id)
                .where(
                    ExpenseRecognitionEntry.venue_id == int(venue.id),
                    ExpenseRecognitionEntry.recognition_date == boundary_date,
                    ExpenseRecognitionEntry.shift_slot == "NIGHT",
                    Expense.status == "CONFIRMED",
                )
            )
            .scalars()
            .all()
        )
        _assert(night_expense_minor > 0, "Common expense has no NIGHT share")
        _assert(
            int(finance["expense_minor"]) == expected_night_expense_minor,
            "Night expense summary does not match recognition entries",
        )
        _assert(finance["slot_profit_available"] is True, "Slot profit must be available after cost allocation")
        expected_profit_minor = (
            revenue_units * 100
            - expected_night_expense_minor
            - int(finance["payroll_minor"])
            + int(finance["adjustments_minor"])
            + int(finance["refunds_minor"])
        )
        _assert(
            int(finance["profit_minor"]) == expected_profit_minor,
            "Night profit did not apply its full allocated cost formula",
        )

        salary = build_member_day_breakdown(
            db=db,
            member_user_id=int(staff.id),
            venue_id=int(venue.id),
            target_date=boundary_date,
            shift_slot="NIGHT",
        )
        _assert(
            int(salary["context"]["minutes_total"]) == 360,
            "Night payroll context did not keep overnight duration",
        )
        _assert(
            int(salary["summary"]["tips_minor"]) == tips_units * 100,
            "Night tips were not allocated to the assigned staff member",
        )

        notification_jobs = (
            db.execute(
                select(NotificationJob)
                .where(NotificationJob.payload_json.contains(boundary_date.isoformat()))
                .order_by(NotificationJob.id.asc())
            )
            .scalars()
            .all()
        )
        matching_jobs = []
        for job in notification_jobs:
            try:
                payload = json.loads(job.payload_json or "{}")
            except Exception:
                continue
            if payload.get("target_date") == boundary_date.isoformat() and payload.get("shift_slot") == "NIGHT":
                matching_jobs.append(job)
        _assert(len(matching_jobs) >= 3, "Night close did not enqueue all three notification jobs")

        report_rows, value_rows = venue_revenue_exports._build_revenue_export_details(
            db=db,
            venue_id=int(venue.id),
            period_start=boundary_date,
            period_end=boundary_date,
        )
        matching_report_rows = [item for item in report_rows if int(item["report_id"]) == report_id]
        _assert(
            len(matching_report_rows) == 1 and matching_report_rows[0]["shift_slot"] == "NIGHT",
            "Revenue export details lost the night report",
        )
        xlsx_size = _build_and_assert_xlsx(
            venue_name=venue.name,
            report_id=report_id,
            report_rows=report_rows,
            value_rows=value_rows,
        )

        _assert_invalid_slot_rejected(db, shift_id=shift_id)

        reopen_result = venue_reports.reopen_daily_report(
            venue_id=int(venue.id),
            report_date=boundary_date,
            shift_slot="NIGHT",
            db=db,
            user=owner,
        )
        _assert(reopen_result["status"] == "DRAFT", "Night report did not reopen")
        db.expire_all()
        report = db.get(DailyReport, report_id)
        _assert(report is not None and report.status == "DRAFT", "Reopened report was not persisted")
        revenue_entry_count = int(
            db.execute(
                select(FinanceEntry.id).where(
                    FinanceEntry.source_type == "daily_report",
                    FinanceEntry.source_id == report_id,
                )
            )
            .scalars()
            .first()
            is not None
        )
        _assert(revenue_entry_count == 0, "Reopen left a revenue ledger entry")
        tip_allocation = db.execute(
            select(DailyReportTipAllocation.id).where(DailyReportTipAllocation.report_id == report_id)
        ).scalar_one_or_none()
        _assert(tip_allocation is None, "Reopen left a night tip allocation")

        return {
            "ok": True,
            "venue_id": int(venue.id),
            "shift_id": shift_id,
            "report_id": report_id,
            "common_expense_id": int(common_expense.id),
            "expense_recognition_rows": len(common_expense_rows),
            "night_expense_minor": night_expense_minor,
            "night_expense_total_minor": expected_night_expense_minor,
            "night_payroll_minor": int(finance["payroll_minor"]),
            "night_profit_minor": int(finance["profit_minor"]),
            "shift_date": boundary_date.isoformat(),
            "interval": "22:00–04:00 (+1 день)",
            "duration_minutes": 360,
            "notification_jobs": len(matching_jobs),
            "xlsx_bytes": xlsx_size,
            "final_report_status": report.status,
        }


def main() -> int:
    result = verify_night_shift_boundary()
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
