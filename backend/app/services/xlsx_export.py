from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDEFF5")
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
CURRENCY_FORMAT = '#,##0.00'
INTEGER_FORMAT = '#,##0'
DATETIME_FORMAT = 'yyyy-mm-dd hh:mm'
DATE_FORMAT = 'yyyy-mm-dd'


def _auto_width(ws, col_idx: int, values: Iterable[Any], min_w: int = 8, max_w: int = 60) -> None:
    mx = 0
    for v in values:
        try:
            mx = max(mx, len(str(v)))
        except Exception:
            pass
    w = max(min_w, min(max_w, mx + 2))
    ws.column_dimensions[get_column_letter(col_idx)].width = w



def _minor_to_major(value: Any) -> float:
    try:
        return int(value or 0) / 100.0
    except Exception:
        return 0.0



def _write_title(ws, title: str) -> None:
    ws.append([title])
    ws[1][0].font = TITLE_FONT
    ws[1][0].alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])



def _write_key_values(ws, rows: list[tuple[str, Any]]) -> None:
    for key, value in rows:
        ws.append([key, value])
    if rows:
        for row in ws.iter_rows(min_row=3, max_row=2 + len(rows), min_col=1, max_col=2):
            row[0].font = HEADER_FONT
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    _apply_auto_width(ws)



def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    currency_cols: set[int] | None = None,
    integer_cols: set[int] | None = None,
    datetime_cols: set[int] | None = None,
    date_cols: set[int] | None = None,
) -> None:
    start_row = ws.max_row + 1
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in rows:
        ws.append(row)

    end_row = ws.max_row
    if end_row >= header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{end_row}"

    currency_cols = currency_cols or set()
    integer_cols = integer_cols or set()
    datetime_cols = datetime_cols or set()
    date_cols = date_cols or set()

    for row in ws.iter_rows(min_row=header_row + 1, max_row=end_row, min_col=1, max_col=len(headers)):
        for idx, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if idx in currency_cols and isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FORMAT
            elif idx in integer_cols and isinstance(cell.value, (int, float)):
                cell.number_format = INTEGER_FORMAT
            elif idx in datetime_cols and cell.value is not None:
                cell.number_format = DATETIME_FORMAT
            elif idx in date_cols and cell.value is not None:
                cell.number_format = DATE_FORMAT

    _apply_auto_width(ws)



def _apply_auto_width(ws) -> None:
    for idx, col_cells in enumerate(ws.columns, start=1):
        values = [cell.value for cell in col_cells if cell.value is not None]
        _auto_width(ws, idx, values, min_w=10, max_w=48)



def _finalize_workbook(wb: Workbook) -> bytes:
    out = BytesIO()
    wb.save(out)
    return out.getvalue()



def build_revenue_xlsx(
    *,
    month: str,
    mode: str,
    venue_name: str,
    rows: list[dict[str, Any]],
    total: int,
    closed_reports: int,
    report_rows: list[dict[str, Any]] | None = None,
    value_rows: list[dict[str, Any]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    mode_title = "Оплаты" if str(mode).upper() == "PAYMENTS" else "Департаменты"
    _write_title(ws, f"Выручка · {venue_name}")
    _write_key_values(
        ws,
        [
            ("Период", month),
            ("Режим агрегации", mode_title),
            ("Закрытых отчётов", int(closed_reports or 0)),
            ("Итого, ₽", float(int(total or 0))),
        ],
    )
    ws[6][1].number_format = CURRENCY_FORMAT

    breakdown_ws = wb.create_sheet("Агрегация")
    _write_title(breakdown_ws, f"Агрегация выручки · {mode_title}")
    _write_table(
        breakdown_ws,
        ["Код", "Категория", "Сумма, ₽"],
        [[r.get("code"), r.get("title") or "—", float(int(r.get("amount") or 0))] for r in rows],
        currency_cols={3},
    )

    reports_ws = wb.create_sheet("Отчёты")
    report_rows = report_rows or []
    _write_title(reports_ws, "Закрытые отчёты за период")
    _write_table(
        reports_ws,
        [
            "Дата",
            "Report ID",
            "Статус",
            "Выручка итого, ₽",
            "Оплаты итого, ₽",
            "Департаменты итого, ₽",
            "Расхождение, ₽",
            "Чаевые, ₽",
            "Комментарий",
            "Закрыт",
            "Кем закрыт",
        ],
        [
            [
                item.get("date"),
                item.get("report_id"),
                item.get("status"),
                _minor_to_major(item.get("revenue_total_minor")),
                _minor_to_major(item.get("payments_total_minor")),
                _minor_to_major(item.get("departments_total_minor")),
                _minor_to_major(item.get("discrepancy_minor")),
                _minor_to_major(item.get("tips_total_minor")),
                item.get("comment"),
                item.get("closed_at"),
                item.get("closed_by"),
            ]
            for item in report_rows
        ],
        currency_cols={4, 5, 6, 7, 8},
        date_cols={1},
        datetime_cols={10},
    )

    values_ws = wb.create_sheet("Значения")
    value_rows = value_rows or []
    _write_title(values_ws, "Значения отчётов")
    _write_table(
        values_ws,
        ["Дата", "Report ID", "Тип", "Код", "Название", "Значение"],
        [
            [
                item.get("date"),
                item.get("report_id"),
                item.get("kind"),
                item.get("code"),
                item.get("title"),
                item.get("value_numeric"),
            ]
            for item in value_rows
        ],
        integer_cols={6},
        date_cols={1},
    )

    return _finalize_workbook(wb)



def build_revenue_csv(
    *,
    month: str,
    mode: str,
    venue_name: str,
    rows: list[dict[str, Any]],
    total: int,
    closed_reports: int,
    delimiter: str = ";",
) -> str:
    lines: list[list[str]] = []
    lines.append(["venue", venue_name])
    lines.append(["month", month])
    lines.append(["mode", mode])
    lines.append(["closed_reports", str(closed_reports)])
    lines.append([])
    lines.append(["Категория", "Сумма"])
    for r in rows:
        lines.append([str(r.get("title") or "—"), str(int(r.get("amount") or 0))])
    lines.append(["ИТОГО", str(int(total))])

    def esc(s: str) -> str:
        s = str(s)
        return '"' + s.replace('"', '""') + '"'

    return "\n".join(delimiter.join(esc(x) for x in row) for row in lines)



def build_expenses_xlsx(
    *,
    month: str,
    venue_name: str,
    rows: list[dict[str, Any]],
    total_minor: int,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    _write_title(ws, f"Расходы · {venue_name}")
    _write_key_values(
        ws,
        [
            ("Месяц признания", month),
            ("Записей", len(rows)),
            ("Признано в месяце, ₽", _minor_to_major(total_minor)),
            ("Полная сумма документов, ₽", _minor_to_major(sum(int(r.get("amount_minor") or 0) for r in rows))),
        ],
    )
    ws[4][1].number_format = INTEGER_FORMAT
    ws[5][1].number_format = CURRENCY_FORMAT
    ws[6][1].number_format = CURRENCY_FORMAT

    docs_ws = wb.create_sheet("Документы")
    _write_title(docs_ws, "Список расходов")
    _write_table(
        docs_ws,
        [
            "ID",
            "Дата",
            "Статус",
            "Категория",
            "Поставщик",
            "Оплачено через",
            "Полная сумма, ₽",
            "Признано в месяце, ₽",
            "Месяцев распределения",
            "Сгенерировано на месяц",
            "Recurring rule ID",
            "Комментарий",
            "Создано",
            "Обновлено",
        ],
        [
            [
                item.get("id"),
                item.get("expense_date"),
                item.get("status"),
                (item.get("category") or {}).get("title"),
                (item.get("supplier") or {}).get("title"),
                (item.get("payment_method") or {}).get("title"),
                _minor_to_major(item.get("amount_minor")),
                _minor_to_major(item.get("recognized_amount_minor_for_month")),
                item.get("spread_months"),
                item.get("generated_for_month"),
                item.get("recurring_rule_id"),
                item.get("comment"),
                item.get("created_at"),
                item.get("updated_at"),
            ]
            for item in rows
        ],
        currency_cols={7, 8},
        integer_cols={1, 9, 11},
        date_cols={2, 10},
        datetime_cols={13, 14},
    )

    alloc_ws = wb.create_sheet("Аллокации")
    _write_title(alloc_ws, "Распределение по месяцам")
    alloc_rows: list[list[Any]] = []
    for item in rows:
        category_title = (item.get("category") or {}).get("title")
        for alloc in item.get("allocations") or []:
            alloc_rows.append(
                [
                    item.get("id"),
                    item.get("expense_date"),
                    category_title,
                    alloc.get("month"),
                    _minor_to_major(alloc.get("amount_minor")),
                    "Да" if alloc.get("month") == f"{month}-01" else "Нет",
                ]
            )
    _write_table(
        alloc_ws,
        ["Expense ID", "Дата расхода", "Категория", "Месяц признания", "Сумма, ₽", "В месяце экспорта"],
        alloc_rows,
        currency_cols={5},
        integer_cols={1},
        date_cols={2, 4},
    )

    return _finalize_workbook(wb)



def build_monthly_summary_xlsx(
    *,
    month: str | None,
    period_start: Any | None,
    period_end: Any | None,
    venue_name: str,
    payments_summary: dict[str, Any],
    departments_summary: dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    period_label = month or ""
    if not period_label:
        if period_start and period_end:
            period_label = f"{period_start.isoformat()} — {period_end.isoformat()}"
        elif period_start:
            period_label = str(period_start)
        else:
            period_label = "—"
    _write_title(ws, f"Сводка · {venue_name}")
    _write_key_values(
        ws,
        [
            ("Период", period_label),
            ("Выручка, ₽", _minor_to_major(payments_summary.get("revenue_minor"))),
            ("Расходы без ФОТ, ₽", _minor_to_major(payments_summary.get("expense_without_payroll_minor"))),
            ("ФОТ, ₽", _minor_to_major(payments_summary.get("payroll_minor"))),
            ("Всего затрат, ₽", _minor_to_major(payments_summary.get("total_cost_minor"))),
            ("Прибыль, ₽", _minor_to_major(payments_summary.get("profit_minor"))),
            ("Маржинальность, %", (payments_summary.get("margin_bps") or 0) / 100 if payments_summary.get("margin_bps") is not None else None),
            ("Черновиков расходов", int(payments_summary.get("draft_expense_count") or 0)),
            ("Черновики на сумму, ₽", _minor_to_major(payments_summary.get("draft_expense_total_minor"))),
        ],
    )
    for row_idx in (4, 5, 6, 7, 8, 11):
        ws[row_idx][1].number_format = CURRENCY_FORMAT
    ws[9][1].number_format = '0.00'
    ws[10][1].number_format = INTEGER_FORMAT

    payments_ws = wb.create_sheet("Выручка по оплатам")
    _write_title(payments_ws, "Выручка по способам оплат")
    _write_table(
        payments_ws,
        ["Код", "Статья", "Сумма, ₽"],
        [[row.get("code"), row.get("title"), _minor_to_major(row.get("amount_minor"))] for row in (payments_summary.get("revenue_breakdown") or [])],
        currency_cols={3},
    )

    dept_ws = wb.create_sheet("Выручка по департаментам")
    _write_title(dept_ws, "Выручка по департаментам")
    _write_table(
        dept_ws,
        ["Код", "Департамент", "Сумма, ₽"],
        [[row.get("code"), row.get("title"), _minor_to_major(row.get("amount_minor"))] for row in (departments_summary.get("revenue_breakdown") or [])],
        currency_cols={3},
    )

    expenses_ws = wb.create_sheet("Расходы по категориям")
    _write_title(expenses_ws, "Расходы по категориям")
    _write_table(
        expenses_ws,
        ["Код", "Категория", "Сумма, ₽"],
        [[row.get("code"), row.get("title"), _minor_to_major(row.get("amount_minor"))] for row in (payments_summary.get("expense_categories") or [])],
        currency_cols={3},
    )

    balances_ws = wb.create_sheet("Балансы оплат")
    _write_title(balances_ws, "Движение по способам оплат")
    _write_table(
        balances_ws,
        ["Код", "Способ оплаты", "Поступления, ₽", "Списания, ₽", "Баланс, ₽"],
        [
            [
                row.get("code"),
                row.get("title"),
                _minor_to_major(row.get("inflow_minor")),
                _minor_to_major(row.get("outflow_minor")),
                _minor_to_major(row.get("balance_minor")),
            ]
            for row in (payments_summary.get("payment_method_balances") or [])
        ],
        currency_cols={3, 4, 5},
    )

    return _finalize_workbook(wb)



def build_payroll_xlsx(
    *,
    period_label: str,
    venue_name: str,
    payload: dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    run = payload.get("run") or {}
    latest_recalculation = payload.get("latest_recalculation") or {}
    lines = payload.get("lines") or []
    _write_title(ws, f"Начисления · {venue_name}")
    _write_key_values(
        ws,
        [
            ("Период", period_label),
            ("Строк начислений", int(payload.get("lines_count") or 0)),
            ("Итого, ₽", _minor_to_major(payload.get("total_amount_minor"))),
            ("Рассчитано", run.get("calculated_at")),
            ("Последний перерасчёт", latest_recalculation.get("created_at")),
            ("Причина перерасчёта", latest_recalculation.get("trigger_reason")),
            ("Run ID", run.get("id")),
        ],
    )
    ws[4][1].number_format = INTEGER_FORMAT
    ws[5][1].number_format = CURRENCY_FORMAT
    ws[9][1].number_format = INTEGER_FORMAT

    lines_ws = wb.create_sheet("Начисления")
    _write_title(lines_ws, "Строки начислений")
    _write_table(
        lines_ws,
        ["Сотрудник", "Username", "Профиль", "Сумма, ₽", "Часы", "Смены", "Отработано дней"],
        [
            [
                (line.get("member") or {}).get("short_name") or (line.get("member") or {}).get("full_name") or f"user #{line.get('member_user_id')}",
                (line.get("member") or {}).get("tg_username"),
                line.get("pay_profile_title"),
                _minor_to_major(line.get("amount_minor")),
                ((line.get("breakdown") or {}).get("metrics") or {}).get("hours_total"),
                ((line.get("breakdown") or {}).get("metrics") or {}).get("shifts_count"),
                ((line.get("breakdown") or {}).get("metrics") or {}).get("worked_dates_count"),
            ]
            for line in lines
        ],
        currency_cols={4},
        integer_cols={5, 6, 7},
    )

    comp_ws = wb.create_sheet("Разбор")
    _write_title(comp_ws, "Детализация компонентов")
    component_rows: list[list[Any]] = []
    percent_day_rows: list[list[Any]] = []
    for line in lines:
        member = (line.get("member") or {}).get("short_name") or (line.get("member") or {}).get("full_name") or f"user #{line.get('member_user_id')}"
        profile = line.get("pay_profile_title")
        for comp in (((line.get("breakdown") or {}).get("components")) or []):
            matched_step = comp.get("matched_step") or {}
            snapshot = comp.get("calculation_snapshot") if isinstance(comp.get("calculation_snapshot"), dict) else {}
            base_scope_title = snapshot.get("base_scope_title") or comp.get("base_scope_title")
            boost_source_title = snapshot.get("boost_source_title") or comp.get("boost_source_title")
            boost_mode_title = snapshot.get("boost_recalc_mode_title") or comp.get("boost_recalc_mode_title")
            boost_mode_effective = snapshot.get("boost_recalc_mode_effective") or comp.get("boost_recalc_mode_effective")
            boost_target_minor = snapshot.get("boost_target_minor") if snapshot.get("boost_target_minor") is not None else comp.get("boost_target_minor")
            boost_actual_minor = snapshot.get("boost_actual_minor") if snapshot.get("boost_actual_minor") is not None else comp.get("boost_actual_minor")
            boost_target_value = snapshot.get("boost_target_value") if snapshot.get("boost_target_value") is not None else comp.get("boost_target_value")
            boost_actual_value = snapshot.get("boost_actual_value") if snapshot.get("boost_actual_value") is not None else comp.get("boost_actual_value")
            component_rows.append(
                [
                    member,
                    profile,
                    comp.get("title") or comp.get("component_type"),
                    comp.get("component_type"),
                    _minor_to_major(comp.get("amount_minor")),
                    _minor_to_major(comp.get("base_amount_minor")) if comp.get("base_amount_minor") is not None else None,
                    base_scope_title,
                    (comp.get("regular_percent_bps") or 0) / 100 if comp.get("regular_percent_bps") is not None else None,
                    (comp.get("percent_bps") or 0) / 100 if comp.get("percent_bps") is not None else None,
                    boost_source_title,
                    _minor_to_major(boost_target_minor) if boost_target_minor is not None else None,
                    _minor_to_major(boost_actual_minor) if boost_actual_minor is not None else None,
                    boost_target_value,
                    boost_actual_value,
                    (comp.get("boost_percent_bps") or 0) / 100 if comp.get("boost_percent_bps") is not None else None,
                    boost_mode_title,
                    boost_mode_effective,
                    _minor_to_major(comp.get("minimum_guarantee_minor")) if comp.get("minimum_guarantee_minor") is not None else None,
                    _minor_to_major(comp.get("maximum_cap_minor")) if comp.get("maximum_cap_minor") is not None else None,
                    comp.get("department_title"),
                    comp.get("boost_department_title"),
                    comp.get("boost_kpi_metric_title") or comp.get("kpi_metric_title"),
                    "Да" if comp.get("boost_applied") else "Нет",
                    "Да" if comp.get("minimum_applied") else "Нет",
                    "Да" if comp.get("maximum_applied") else "Нет",
                    matched_step.get("threshold_value"),
                    matched_step.get("bonus_minor") / 100.0 if matched_step.get("bonus_minor") is not None else None,
                ]
            )
            for day_row in (snapshot.get("day_rows") or comp.get("day_rows") or []):
                if not isinstance(day_row, dict):
                    continue
                percent_day_rows.append(
                    [
                        member,
                        profile,
                        comp.get("title") or comp.get("component_type"),
                        boost_source_title,
                        day_row.get("date"),
                        _minor_to_major(day_row.get("base_amount_minor")) if day_row.get("base_amount_minor") is not None else None,
                        _minor_to_major(day_row.get("target_amount_minor")) if day_row.get("target_amount_minor") is not None else None,
                        _minor_to_major(day_row.get("actual_amount_minor")) if day_row.get("actual_amount_minor") is not None else None,
                        (day_row.get("percent_bps") or 0) / 100 if day_row.get("percent_bps") is not None else None,
                        _minor_to_major(day_row.get("amount_minor")) if day_row.get("amount_minor") is not None else None,
                        "Да" if day_row.get("boost_applied") else "Нет",
                        boost_mode_effective,
                    ]
                )
    _write_table(
        comp_ws,
        [
            "Сотрудник",
            "Профиль",
            "Компонент",
            "Тип",
            "Сумма, ₽",
            "База, ₽",
            "База расчёта",
            "Обычный %, %",
            "Применённый %, %",
            "Условие",
            "Цель, ₽",
            "Факт, ₽",
            "Цель KPI",
            "Факт KPI",
            "Повышенный %, %",
            "Режим",
            "Эффективный режим",
            "Мин. гарантия, ₽",
            "Максимум, ₽",
            "Департамент",
            "Департамент условия",
            "KPI",
            "Boost",
            "Мин. гарантия",
            "Потолок",
            "Сработавшая ступень",
            "Бонус ступени, ₽",
        ],
        component_rows,
        currency_cols={5, 6, 11, 12, 18, 19, 27},
    )

    if percent_day_rows:
        day_ws = wb.create_sheet("Проценты по дням")
        _write_title(day_ws, "Посуточная детализация процентных компонентов")
        _write_table(
            day_ws,
            [
                "Сотрудник",
                "Профиль",
                "Компонент",
                "Условие",
                "Дата",
                "База, ₽",
                "Цель, ₽",
                "Факт, ₽",
                "Применённый %, %",
                "Сумма, ₽",
                "Boost",
                "Режим",
            ],
            percent_day_rows,
            currency_cols={6, 7, 8, 10},
        )

    return _finalize_workbook(wb)

    return _finalize_workbook(wb)
