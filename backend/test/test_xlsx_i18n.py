from __future__ import annotations

from datetime import date
from io import BytesIO
import re
from unittest import TestCase

from openpyxl import load_workbook

from app.services.xlsx_export import (
    build_billing_reconciliation_xlsx,
    build_billing_transactions_xlsx,
    build_expenses_xlsx,
    build_finance_ledger_xlsx,
    build_monthly_summary_xlsx,
    build_payroll_xlsx,
    build_revenue_csv,
    build_revenue_xlsx,
)


_CYRILLIC = re.compile(r"[А-Яа-яЁё]")


def _workbook_strings(payload: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    values = [sheet.title for sheet in workbook.worksheets]
    for sheet in workbook.worksheets:
        values.extend(str(cell.value) for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str))
    workbook.close()
    return values


class XlsxI18nTests(TestCase):
    def test_english_xlsx_exports_have_no_russian_system_labels(self) -> None:
        payloads = [
            build_finance_ledger_xlsx(
                venue_name="North Bar",
                period_start=date(2026, 8, 1),
                period_end=date(2026, 8, 31),
                rows=[],
                filters=[("Направление", "Приход")],
                locale="en",
            ),
            build_revenue_xlsx(
                month="2026-08",
                mode="PAYMENTS",
                venue_name="North Bar",
                rows=[],
                total=0,
                closed_reports=0,
                locale="en",
            ),
            build_expenses_xlsx(
                month="2026-08",
                venue_name="North Bar",
                rows=[],
                total_minor=0,
                locale="en",
            ),
            build_monthly_summary_xlsx(
                month="2026-08",
                period_start=date(2026, 8, 1),
                period_end=date(2026, 8, 31),
                venue_name="North Bar",
                payments_summary={},
                departments_summary={},
                locale="en",
            ),
            build_payroll_xlsx(
                period_label="2026-08",
                venue_name="North Bar",
                payload={},
                locale="en",
            ),
            build_billing_transactions_xlsx(
                title="Axelio · Подписка · North Bar",
                rows=[],
                filters=[("Заведение", "North Bar")],
                locale="en",
            ),
            build_billing_reconciliation_xlsx(
                title="Axelio · Billing reconciliation",
                rows=[],
                filters=[("Status", "Все")],
                locale="en",
            ),
        ]

        untranslated = [
            value for payload in payloads for value in _workbook_strings(payload) if _CYRILLIC.search(value)
        ]
        self.assertEqual(untranslated, [])

    def test_english_revenue_csv_headers_are_localized(self) -> None:
        csv_text = build_revenue_csv(
            month="2026-08",
            mode="PAYMENTS",
            venue_name="North Bar",
            rows=[],
            total=0,
            closed_reports=0,
            locale="en",
        )

        self.assertIn("Category", csv_text)
        self.assertIn("Amount", csv_text)
        self.assertIn("TOTAL", csv_text)
        self.assertIsNone(_CYRILLIC.search(csv_text))
