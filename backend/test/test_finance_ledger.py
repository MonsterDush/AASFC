from __future__ import annotations

from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest import TestCase
from unittest.mock import patch

from openpyxl import load_workbook

from app.routers.venue_ledger import (
    _finance_entries_statement,
    _load_finance_entry_analytics,
    _load_finance_entry_payload,
    _serialize_finance_entry,
)
from app.services.finance.ledger import create_finance_entry
from app.services.finance.reconciliation import build_finance_reconciliation
from app.services.xlsx_export import build_finance_ledger_xlsx


class _FakeSession:
    def __init__(self):
        self.added = []

    def add(self, obj):
        self.added.append(obj)


class _RowsResult:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return list(self.rows)

    def one(self):
        if len(self.rows) != 1:
            raise AssertionError(f"expected one row, got {len(self.rows)}")
        return self.rows[0]

    def scalar_one_or_none(self):
        if not self.rows:
            return None
        if len(self.rows) != 1:
            raise AssertionError(f"expected zero or one row, got {len(self.rows)}")
        row = self.rows[0]
        return row[0] if isinstance(row, tuple) else row


class _ReconciliationSession:
    def __init__(self, result_rows):
        self.result_rows = list(result_rows)

    def execute(self, _statement):
        return _RowsResult(self.result_rows.pop(0))


class _CaptureSession:
    def __init__(self, result_rows):
        self.result_rows = list(result_rows)
        self.statements = []

    def execute(self, statement):
        self.statements.append(statement)
        return _RowsResult(self.result_rows.pop(0))


class FinanceLedgerTests(TestCase):
    def test_serialize_legacy_report_entry_adds_shift_slot_from_source(self):
        entry = SimpleNamespace(
            id=4,
            venue_id=12,
            entry_date=date(2026, 3, 10),
            amount_minor=12345,
            direction="INCOME",
            kind="REVENUE",
            source_type="daily_report",
            source_id=77,
            meta_json={"report_date": "2026-03-10"},
            payment_method=None,
            department=None,
            created_at=None,
        )

        payload = _serialize_finance_entry(entry, report_shift_slot="NIGHT")

        self.assertEqual(payload["meta_json"]["shift_slot"], "NIGHT")
        self.assertEqual(entry.meta_json, {"report_date": "2026-03-10"})

    def test_create_finance_entry_uses_kopecks_and_absolute_amount(self):
        db = _FakeSession()

        entry = create_finance_entry(
            db=db,
            venue_id=12,
            entry_date=date(2026, 3, 10),
            amount_minor=12345,
            direction="income",
            kind="revenue",
            source_type="daily_report",
            source_id=77,
            meta_json={"report_date": "2026-03-10"},
        )

        self.assertEqual(entry.amount_minor, 12345)
        self.assertEqual(entry.direction, "INCOME")
        self.assertEqual(entry.kind, "REVENUE")
        self.assertEqual(entry.source_type, "daily_report")
        self.assertEqual(entry.source_id, 77)
        self.assertEqual(entry.meta_json, {"report_date": "2026-03-10"})
        self.assertEqual(len(db.added), 1)

    def test_create_finance_entry_rejects_non_int_amount(self):
        db = _FakeSession()

        with self.assertRaisesRegex(ValueError, "amount_minor must be int"):
            create_finance_entry(
                db=db,
                venue_id=12,
                entry_date=date(2026, 3, 10),
                amount_minor=12.34,
                direction="income",
                kind="revenue",
                source_type="daily_report",
            )

    def test_create_finance_entry_rejects_negative_amount(self):
        db = _FakeSession()

        with self.assertRaisesRegex(ValueError, "amount_minor must be non-negative"):
            create_finance_entry(
                db=db,
                venue_id=12,
                entry_date=date(2026, 3, 10),
                amount_minor=-100,
                direction="expense",
                kind="expense",
                source_type="expense",
            )

    def test_reconciliation_uses_source_basis_for_expenses(self):
        db = _ReconciliationSession(
            [
                [("daily_report", 10, 10_000, 1)],
                [(10, date(2026, 7, 3), 100)],
                [("expense", 20, 5_000, 1)],
                [(20, date(2026, 7, 4), 5_000)],
                [("payroll_run", 30, 7_000, 1)],
                [],
                [(30, date(2026, 7, 1), 7_000)],
            ]
        )
        with patch(
            "app.services.finance.reconciliation.get_finance_summary",
            return_value={
                "revenue_minor": 10_000,
                "expense_without_payroll_minor": 3_000,
                "payroll_minor": 7_000,
            },
        ):
            payload = build_finance_reconciliation(db=db, venue_id=5, month="2026-07")

        self.assertEqual(payload["status"], "OK")
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertTrue(checks["revenue"]["comparable_to_summary"])
        self.assertFalse(checks["expense"]["comparable_to_summary"])
        self.assertEqual(checks["expense"]["status"], "OK")
        self.assertEqual(checks["expense"]["summary_minor"], 3_000)
        self.assertEqual(checks["expense"]["source_ledger_minor"], 5_000)
        self.assertTrue(checks["payroll"]["comparable_to_summary"])

    def test_reconciliation_reports_problem_sources_without_false_partial_payroll_warning(self):
        db = _ReconciliationSession(
            [
                [("daily_report", 10, 9_000, 1)],
                [(10, date(2026, 7, 10), 100)],
                [],
                [(20, date(2026, 7, 12), 5_000)],
                [],
                [],
            ]
        )
        with patch(
            "app.services.finance.reconciliation.get_finance_summary",
            return_value={
                "revenue_minor": 10_000,
                "expense_without_payroll_minor": 5_000,
                "payroll_minor": 2_000,
            },
        ):
            payload = build_finance_reconciliation(
                db=db,
                venue_id=5,
                date_from=date(2026, 7, 10),
                date_to=date(2026, 7, 20),
            )

        self.assertEqual(payload["status"], "WARNING")
        self.assertEqual(payload["issue_count"], 2)
        self.assertEqual({item["reason"] for item in payload["issues"]}, {"AMOUNT_MISMATCH", "MISSING_LEDGER_ENTRY"})
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["payroll"]["status"], "INFO")
        self.assertFalse(checks["payroll"]["comparable_to_summary"])

    def test_reconciliation_compares_configured_payroll_payouts_instead_of_accruals(self):
        db = _ReconciliationSession(
            [
                [],
                [],
                [],
                [],
                [("payroll_expense", 40, 7_000, 1)],
                [(1,)],
                [(40, date(2026, 7, 20), 7_000)],
            ]
        )
        with patch(
            "app.services.finance.reconciliation.get_finance_summary",
            return_value={
                "revenue_minor": 0,
                "expense_without_payroll_minor": 0,
                "payroll_minor": 9_000,
            },
        ):
            payload = build_finance_reconciliation(db=db, venue_id=5, month="2026-07")

        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["payroll"]["status"], "OK")
        self.assertFalse(checks["payroll"]["comparable_to_summary"])
        self.assertEqual(checks["payroll"]["source_type"], "payroll_expense")
        self.assertEqual(checks["payroll"]["source_expected_minor"], 7_000)

    def test_finance_ledger_xlsx_keeps_numeric_money_and_filterable_rows(self):
        data = build_finance_ledger_xlsx(
            venue_name="Тест",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 31),
            filters=[("Направление", "Приход")],
            rows=[
                {
                    "id": 8,
                    "entry_date": "2026-07-03",
                    "amount_minor": 12_345,
                    "direction": "INCOME",
                    "kind": "REVENUE",
                    "source_type": "daily_report",
                    "source_id": 10,
                    "payment_method": {"title": "Карта"},
                    "department": None,
                    "meta_json": {"shift_slot": "NIGHT"},
                    "created_at": "2026-07-03T04:30:00+03:00",
                }
            ],
        )
        workbook = load_workbook(BytesIO(data), data_only=False)
        worksheet = workbook["Операции"]
        header_row = next(row[0].row for row in worksheet.iter_rows() if row[0].value == "Дата")
        self.assertEqual(worksheet.cell(header_row + 1, 1).value.date(), date(2026, 7, 3))
        self.assertEqual(worksheet.cell(header_row + 1, 4).value, 123.45)
        self.assertEqual(worksheet.cell(header_row + 1, 4).number_format, "#,##0.00")
        self.assertEqual(worksheet.freeze_panes, f"A{header_row + 1}")
        self.assertEqual(worksheet.auto_filter.ref, f"A{header_row}:M{header_row + 1}")
        self.assertEqual(worksheet.sheet_view.showGridLines, False)

    def test_shared_ledger_statement_applies_every_export_filter(self):
        statement = _finance_entries_statement(
            venue_id=5,
            month="2026-07",
            date_from=None,
            date_to=None,
            payment_method_id=3,
            direction="income",
            kind="revenue",
            source_type="daily_report",
        )
        sql = str(statement.compile(compile_kwargs={"literal_binds": True}))

        self.assertIn("finance_entries.entry_date >= '2026-07-01'", sql)
        self.assertIn("finance_entries.entry_date <= '2026-07-31'", sql)
        self.assertIn("finance_entries.payment_method_id = 3", sql)
        self.assertIn("finance_entries.direction = 'INCOME'", sql)
        self.assertIn("finance_entries.kind = 'REVENUE'", sql)
        self.assertIn("finance_entries.source_type = 'daily_report'", sql)

    def test_finance_entry_payload_applies_page_limit_and_offset(self):
        db = _CaptureSession([[]])

        payload = _load_finance_entry_payload(
            db,
            venue_id=5,
            month="2026-07",
            date_from=None,
            date_to=None,
            payment_method_id=None,
            direction=None,
            kind=None,
            source_type=None,
            limit=51,
            offset=50,
        )

        sql = str(db.statements[0].compile(compile_kwargs={"literal_binds": True}))
        self.assertEqual(payload, [])
        self.assertIn("LIMIT 51", sql)
        self.assertIn("OFFSET 50", sql)

    def test_finance_entry_analytics_returns_compact_metrics_series_and_structure(self):
        db = _CaptureSession(
            [
                [(150_000, 40_000, 5)],
                [(date(2026, 7, 3), 100_000, 25_000, 3), (date(2026, 7, 4), 50_000, 15_000, 2)],
                [("INCOME", "REVENUE", 150_000, 2), ("EXPENSE", "EXPENSE", 40_000, 3)],
            ]
        )

        payload = _load_finance_entry_analytics(
            db,
            venue_id=5,
            month="2026-07",
            date_from=None,
            date_to=None,
            payment_method_id=None,
            direction=None,
            kind=None,
            source_type=None,
        )

        self.assertEqual(
            payload["metrics"],
            {
                "income_minor": 150_000,
                "expense_minor": 40_000,
                "net_minor": 110_000,
                "count": 5,
            },
        )
        self.assertEqual(payload["daily_series"][0]["count"], 3)
        self.assertEqual(
            payload["structure"][1],
            {
                "direction": "EXPENSE",
                "kind": "EXPENSE",
                "amount_minor": 40_000,
                "count": 3,
            },
        )
